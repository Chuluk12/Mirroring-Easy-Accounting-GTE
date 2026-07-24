"""
Script untuk DELETE kode barang dari database Easy Accounting secara bulk.

⚠️  PERINGATAN KERAS:
    - Script ini LANGSUNG memodifikasi database Easy Accounting (GTE.EASY6)
    - Pastikan Easy Accounting DITUTUP sebelum menjalankan script ini
    - Buat BACKUP database sebelum menjalankan
    - Barang yang sudah punya transaksi sebaiknya TIDAK dihapus
    - Penghapusan tidak bisa di-undo!

Cara pakai:
    1. Isi daftar KODE_BARANG_YANG_DIHAPUS di bawah
    2. Set MODE = "preview" untuk lihat dulu tanpa hapus
    3. Set MODE = "delete" untuk benar-benar hapus
"""

import os
import sys
from pathlib import Path

import fdb
import openpyxl

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

# ─── KONFIGURASI ─────────────────────────────────────────────────────────────

fdb.load_api("C:/Program Files/Firebird/Firebird_3_0/fbclient.dll")

ENV_FILE = Path(__file__).with_name(".env")

def muat_env(file_path):
    """Muat konfigurasi sederhana dari .env tanpa menimpa environment aktif."""
    if not file_path.exists():
        return

    for baris in file_path.read_text(encoding="utf-8").splitlines():
        baris = baris.strip()
        if not baris or baris.startswith("#") or "=" not in baris:
            continue
        nama, nilai = baris.split("=", 1)
        os.environ.setdefault(nama.strip(), nilai.strip())

muat_env(ENV_FILE)

DB_CONFIG = {
    "host": os.getenv("EASY_DB_HOST", "192.168.10.5"),
    "port": int(os.getenv("EASY_DB_PORT", "3999")),
    "database": os.getenv("EASY_DB_PATH", "E:/EASY/GTE.EASY6"),
    "user": os.getenv("EASY_DB_USER", "SYSDBA"),
    "password": os.getenv("EASY_DB_PASSWORD", "NewPassword123"),
}

# ─── ISI KODE BARANG YANG INGIN DIHAPUS ──────────────────────────────────────
# Contoh: ["TEST1", "TEST2", "TEST3"]
# Atau bisa dibaca dari file txt: open("kode_barang.txt").read().splitlines()

FILE_KODE_BARANG = Path(r"C:\Users\HUSNUL\Documents\HasilBarang.xlsx")
KODE_BARANG_DIABAIKAN = {
    "GTE2600AI-6.00-M08",
}

def baca_kode_barang_excel(file_path):
    """Baca kode barang dari kolom pertama, dengan baris pertama sebagai header."""
    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    kode_list = []
    kode_terlihat = set()

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        kode = str(row[0]).strip() if row and row[0] is not None else ""
        if kode and kode not in KODE_BARANG_DIABAIKAN and kode not in kode_terlihat:
            kode_list.append(kode)
            kode_terlihat.add(kode)

    workbook.close()
    return kode_list

KODE_BARANG_YANG_DIHAPUS = baca_kode_barang_excel(FILE_KODE_BARANG)

# "preview" = hanya lihat, tidak hapus
# "delete"  = benar-benar hapus (pastikan sudah preview dulu!)
MODE = "delete"

# ─── SCRIPT ──────────────────────────────────────────────────────────────────

def cek_transaksi(cur, itemno):
    """Cek apakah barang sudah punya transaksi di ITEMHIST."""
    cur.execute("SELECT COUNT(*) FROM ITEMHIST WHERE ITEMNO = ?", (itemno,))
    return cur.fetchone()[0]

def cek_referensi_bom(cur, itemno):
    """Cek jumlah formula BOM yang masih mereferensikan barang."""
    cur.execute("SELECT COUNT(*) FROM BOM WHERE ITEMNO = ?", (itemno,))
    return cur.fetchone()[0]

def preview(con, kode_list):
    cur = con.cursor()
    print("\n" + "="*60)
    print("MODE PREVIEW — tidak ada yang dihapus")
    print("="*60)
    print(f"{'No.':<5} {'Kode Barang':<30} {'Deskripsi':<35} {'Transaksi':<10} {'Status'}")
    print("-"*60)

    bisa_hapus = []
    tidak_bisa = []

    for i, itemno in enumerate(kode_list, 1):
        cur.execute("""
            SELECT ITEMNO, ITEMDESCRIPTION FROM ITEM WHERE ITEMNO = ?
        """, (itemno,))
        row = cur.fetchone()

        if not row:
            print(f"{i:<5} {itemno:<30} {'[TIDAK DITEMUKAN]':<35} {'-':<10} ❌ Tidak ada di database")
            tidak_bisa.append(itemno)
            continue

        jumlah_tx = cek_transaksi(cur, itemno)
        jumlah_bom = cek_referensi_bom(cur, itemno)
        desc = str(row[1] or "").strip()[:33]

        if jumlah_tx > 0:
            print(f"{i:<5} {itemno:<30} {desc:<35} {jumlah_tx:<10} ⚠️  Ada transaksi — SKIP")
            tidak_bisa.append(itemno)
        elif jumlah_bom > 0:
            print(f"{i:<5} {itemno:<30} {desc:<35} {jumlah_tx:<10} Dipakai {jumlah_bom} BOM - SKIP")
            tidak_bisa.append(itemno)
        else:
            print(f"{i:<5} {itemno:<30} {desc:<35} {jumlah_tx:<10} ✅ Bisa dihapus")
            bisa_hapus.append(itemno)

    print("-"*60)
    print(f"Total: {len(kode_list)} | Bisa hapus: {len(bisa_hapus)} | Skip: {len(tidak_bisa)}")
    print("\nUntah menghapus, ganti MODE = 'delete' lalu jalankan ulang.")
    return bisa_hapus

def delete_barang(con, kode_list):
    cur = con.cursor()
    print("\n" + "="*60)
    print("MODE DELETE — menghapus barang dari database")
    print("="*60)

    berhasil = []
    gagal = []

    for itemno in kode_list:
        # Cek ada di database
        cur.execute("SELECT ITEMNO, ITEMDESCRIPTION FROM ITEM WHERE ITEMNO = ?", (itemno,))
        row = cur.fetchone()

        if not row:
            print(f"❌ {itemno} — tidak ditemukan, skip")
            gagal.append(itemno)
            continue

        # Cek transaksi
        jumlah_tx = cek_transaksi(cur, itemno)
        if jumlah_tx > 0:
            print(f"⚠️  {itemno} — ada {jumlah_tx} transaksi, skip (tidak dihapus)")
            gagal.append(itemno)
            continue

        jumlah_bom = cek_referensi_bom(cur, itemno)
        if jumlah_bom > 0:
            print(f"{itemno} - dipakai {jumlah_bom} formula BOM, skip (tidak dihapus)")
            gagal.append(itemno)
            continue

        try:
            # Hapus dari tabel-tabel terkait dulu
            cur.execute("DELETE FROM ITEM_BRANCH WHERE ITEMNO = ?", (itemno,))
            cur.execute("DELETE FROM WAREITEM WHERE ITEMNO = ?", (itemno,))
            cur.execute("DELETE FROM ITEMBALANCE WHERE ITEMNO = ?", (itemno,))
            cur.execute("DELETE FROM SN WHERE ITEMNO = ?", (itemno,))
            cur.execute("DELETE FROM ITEMORDER WHERE ITEMNO = ?", (itemno,))
            cur.execute("DELETE FROM PERSONDATAPRICE WHERE ITEMNO = ?", (itemno,))
            cur.execute("DELETE FROM REPLACEITEM WHERE ITEMNO = ?", (itemno,))
            cur.execute("DELETE FROM REPLACEITEM WHERE RITEMNO = ?", (itemno,))
            # Hapus dari ITEM
            cur.execute("DELETE FROM ITEM WHERE ITEMNO = ?", (itemno,))
            con.commit()
            print(f"✅ {itemno} — berhasil dihapus")
            berhasil.append(itemno)
        except Exception as e:
            con.rollback()
            print(f"❌ {itemno} — ERROR: {e}")
            gagal.append(itemno)

    print("-"*60)
    print(f"Berhasil: {len(berhasil)} | Gagal/Skip: {len(gagal)}")
    if berhasil:
        print(f"\nBarang yang dihapus: {', '.join(berhasil)}")

def main():
    if not KODE_BARANG_YANG_DIHAPUS:
        print("❌ Daftar kode barang kosong! Isi KODE_BARANG_YANG_DIHAPUS dulu.")
        return

    print(f"Menghubungkan ke database Easy Accounting...")
    try:
        con = fdb.connect(**DB_CONFIG)
        print("✅ Koneksi berhasil!")
    except Exception as e:
        print(f"❌ Gagal konek: {e}")
        return

    if MODE == "preview":
        preview(con, KODE_BARANG_YANG_DIHAPUS)
    elif MODE == "delete":
        print("\n⚠️  PERHATIAN: Anda akan menghapus barang dari database Easy Accounting!")
        print("Pastikan Easy Accounting sudah DITUTUP dan database sudah di-BACKUP!")
        konfirmasi = input("\nKetik 'HAPUS' untuk melanjutkan, atau Enter untuk batal: ")
        if konfirmasi.strip() == "HAPUS":
            delete_barang(con, KODE_BARANG_YANG_DIHAPUS)
        else:
            print("Dibatalkan.")
    else:
        print(f"❌ MODE tidak valid: {MODE}. Gunakan 'preview' atau 'delete'.")

    con.close()

if __name__ == "__main__":
    main()
